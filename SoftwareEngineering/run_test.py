from login_test import Login_test
from unittest import TestSuite, TestResult
import pandas as pd
from openpyxl import load_workbook

suite = TestSuite()

data = pd.read_excel('test_code.xlsx')
data

wb = load_workbook('test_code.xlsx')
ws = wb.active

for row in ws['2:6']:
    row_list = []
    for cell in row[1:]:
        row_list.append(cell.value)
    suite.addTest(Login_test('test_login',str(row_list[0]),str(row_list[1]),str(row_list[2])))

# suite.addTest(Login_test(row_list,'admin','123456','登录成功'))
# suite.addTest(Login_test('test_login','admin','1234567','账号密码不正确'))
# suite.addTest(Login_test('test_login','admin','1234','密码长度在6-18之间'))
# suite.addTest(Login_test('test_login','admin','12345678910123123624','密码长度在6-18之间'))
# suite.addTest(Login_test('test_login','administrator','123456','账号密码不正确'))
# print('suite add, successfully')
result = TestResult()
suite.run(result)